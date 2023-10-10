#!/usr/bin/env python
# coding: utf-8

# 1.导入包

# In[26]:


#导入包
import pyodbc
import pandas as pd
import win32com.client as win32
import os
import datetime
import time
import pymssql
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import os
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl import load_workbook
import xlsxwriter
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PIL_Image
import logging as log


# 2.添加log日志

# In[27]:


# 将 log() 函数修改为如下实现，记录日志文件
def log(msg):
    '''记录日志'''
    logs_dir = 'D:/data/auto_PotentialMail/logs'
    if not os.path.exists(logs_dir):
        os.makedirs(logs_dir)
    today_date = datetime.date.today().strftime('%Y%m%d')
    log_path = os.path.join(logs_dir, f'{today_date}.txt')
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with open(log_path, 'a', encoding='utf-8') as f:
        f.write(f'[{timestamp}] {msg}\n')
        
# 修改日志文件保存路径，添加时间戳
logs_dir = 'D:/data/auto_PotentialMail/logs'
if not os.path.exists(logs_dir):
    os.makedirs(logs_dir)
log_path = os.path.join(logs_dir, datetime.date.today().strftime('%Y%m%d') + '.txt')


# 3.配置数据库连接

# In[28]:


# 配置数据库连接
connect = pymssql.connect('GEW-MIS01', 'pbread', 'password',as_dict=True, tds_version='7.0')
log('数据库连接成功')


# 4.读取配置Config参数,在Excel里维护

# In[29]:


# 读取配置文件获取参数
config_file = r"D:\data\auto_PotentialMail\database\config.xlsx"
config_df = pd.read_excel(config_file, sheet_name="Sheet1")
#读取SQL路径
sql_name = "SQL"
sql_path = config_df[config_df['Name'] == sql_name]['Value'].values[0]
#模板路径
template_name = "template"
template_path = config_df[config_df['Name'] == template_name]['Value'].values[0]
#输出子文件路径
output_name = "detail"
output_folder = config_df[config_df['Name'] == output_name]['Value'].values[0]
#输出汇总文件路径
summary_name = "summary"
summary_folder = config_df[config_df['Name'] == summary_name]['Value'].values[0]
#配置邮件信息
outlook = win32.Dispatch('outlook.application')
mail = outlook.CreateItem(0)
#邮件抄送人
cc_name = "cc"
mail.CC = config_df[config_df['Name'] == cc_name]['Value'].values[0]
#邮件发送人
sender_name = "sender"
mail.sender = config_df[config_df['Name'] == sender_name]['Value'].values[0]
#发件邮箱密码
sender_password_name = "sender_password"
sender_password = config_df[config_df['Name'] == sender_password_name]['Value'].values[0]
#邮件内容
body_name = "body"
mail.body = config_df[config_df['Name'] == body_name]['Value'].values[0]
#邮件主题
subject_name = "subject"
mail.subject = config_df[config_df['Name'] == subject_name]['Value'].values[0]
# 获取昨天的日期
yesterday = datetime.date.today() - datetime.timedelta(days=1)
yesterday_str=yesterday.strftime('%Y%m%d')
log('读取config配置')


# In[30]:


#读取SQL脚本
sql_query_Check_Result=''
with open(sql_path,encoding='GB2312')as sql:
    for sql in sql.readlines():
        sql_query_Check_Result=sql_query_Check_Result+sql


# In[31]:


#直接利用pandas读取sql
def get_df_from_db(sql):
    return pd.read_sql_query(sql,connect)
log("读取SQL脚本")


# 4.SQL脚本转换DataFrame：Check_Result

# In[32]:


#把SQL文件转换为DataFrame
Check_Result = get_df_from_db(sql_query_Check_Result)


# 5.关闭数据库连接

# In[33]:


# 关闭数据库连接
connect.close()
log("关闭数据库连接")


# In[34]:


#Check_Result


# #邮箱有误、Sales邮箱地址为空 数量并记录日志

# In[35]:


count = Check_Result[Check_Result['Email'] == '邮箱地址有误'].shape[0]
count_2 = Check_Result[Check_Result['Email'] == 'Sales邮箱地址为空'].shape[0]
wrong_ppo_check_no = Check_Result[Check_Result['Email'] == '邮箱地址有误']['PPO_check_no'].values
wrong_ppo_check_no2 = Check_Result[Check_Result['Email'] == 'Sales邮箱地址为空']['PPO_check_no'].values


# In[36]:


#wrong_ppo_check_no


# In[37]:


if count>=1 or count_2>=1:
    log(f"邮箱地址有误数：{count},ppo_check_no:{wrong_ppo_check_no},Sales邮箱地址为空有误数：{count_2},ppo_check_no:{wrong_ppo_check_no2}")
    


# #删除邮箱地址有误

# In[38]:


Check_Result = Check_Result[~Check_Result['Email'].isin(['邮箱地址有误', 'Sales邮箱地址为空'])].reset_index(drop=True)


# In[39]:


#Check_Result 


# In[40]:


#测试
Check_Result["Email"][0:2]='195537903@qq.com;liangby@esquel.com'
Check_Result["Email"][2:]='liangby@esquel.com;'


# In[41]:


#Check_Result.columns


# In[42]:


#Check_Result.iloc[:, 1:]


# In[43]:


recipient_list = Check_Result['Email'].unique()


# In[44]:


recipient_list


# 6.如果Check_Result不为空，执行发送邮件，Check_Result为空则不发送邮件

#     6.1 批量导出Excel文件
#     6.2 发送邮件

# In[45]:


# 判断 Check_Result 是否为空
if Check_Result.empty:
    log(f'{yesterday_str}评审没有潜在问题反馈')
else:
    #1、Check_Result导出
    summary_file_name = 'Check_Result_' + yesterday.strftime('%Y%m%d') + '.xlsx'  # 导出文件名
    summary_path = os.path.join(summary_folder, summary_file_name)
    #导出Excel
    #Check_Result.to_excel(summary_path, sheet_name='Sheet1', index=False,encoding='GB2312')
    Check_Result.iloc[:, 1:].to_excel(summary_path, sheet_name='Sheet1', index=False,encoding='GB2312')
    log(f'{yesterday_str}评审有潜在问题反馈，共{Check_Result.shape[0]}条，文件导出路径：{summary_path}')

    # 2、将模板的标题行写入Excel文件
    template = pd.read_excel(template_path, sheet_name='Sheet1')
    cols_letters = {}  # 用一个字典来保存每个字段名对应的字母
    for col_num, col_name in enumerate(template.columns):
        col_letter = get_column_letter(col_num+1)
        cols_letters[col_name] = col_letter
    
    # 3、收件人邮箱账号
    #recipient_list = Check_Result['Email'].unique()
    attach_list = []
    #检查邮箱地址是否正确，如不正确，删除
    recipient_list_2 = [recipients for recipients in recipient_list if "@" in recipients]
    drop_recipient = [recipients for recipients in recipient_list if "@" not in recipients]
    #记录日志
    for recipients in drop_recipient:
        log(f"{recipients} 邮箱地址有误") 
    #更新收件人列表
    recipient_list = recipient_list_2

    # 遍历收件人列表
    for i, recipient in enumerate(recipient_list):
        # 获取收件人对应的商品代码
        Customer_code = Check_Result.loc[Check_Result['Email'] == recipient, 'Customer_Code'].iloc[0]

        # 筛选出对应收件人的内容
        recipient_content = Check_Result[Check_Result['Email'] == recipient].reset_index(drop=True)
        #更新从第一列开始取
        recipient_content = recipient_content.iloc[:, 1:]
        #邮件信息
        mail.Subject = f"{subject}——{yesterday.strftime('%Y%m%d')}——测试"
        mail.To = recipient
        # 生成文件名
        Customer_code = recipient_content['Customer_Code'].iloc[0].replace('/', '').replace('.', '') 
        filename = 'Check_Result_{}_{}.xlsx'.format(yesterday.strftime('%Y%m%d'), Customer_code) 
        # 创建日期的输出子文件夹
        result_foder = r"D:\data\auto_PotentialMail\output\detail\{}".format(yesterday_str)
        os.makedirs(result_foder, exist_ok=True)
        
         # 复制模板文件到输出文件夹
        output_path = os.path.join(result_foder, filename)
        os.system(f'copy "{template_path}" "{output_path}"')
        # 将数据写入Excel文件
        writer = pd.ExcelWriter(output_path, engine='openpyxl')
        writer.book = load_workbook(output_path)
        #recipient_content.to_excel(writer, sheet_name='Sheet1', startrow=1,header=False,index=False,encoding='utf-8')
        recipient_content.iloc[:, 1:].to_excel(writer, sheet_name='Sheet1', startrow=1,header=False,index=False,encoding='utf-8')
         # 将数据写入Excel文件
        writer = pd.ExcelWriter(output_path, engine='openpyxl')
        writer.book = load_workbook(output_path)
        sheet = writer.book['Sheet1']
        start_row = 2  # 从第二行开始写入数据
    
        
        #4调整格式
        for index, row in recipient_content.iterrows():
            for col_name, cell_value in row.items():
                # 根据字段名获取列字母
                col_letter = cols_letters[col_name]
                # 获取单元格对象
                cell = sheet[f'{col_letter}{start_row+index}']  
                # 写入数据
                cell.value = cell_value
                # 设置单元格样式，自动换行、垂直顶端对齐、水平左对齐
                new_alignment = Alignment(wrapText=True, vertical='top', horizontal='left')
                cell.alignment = new_alignment
                # 设置边框样式
                if cell_value:
                    border = Border(left=Side(border_style='thin'),
                                    right=Side(border_style='thin'),
                                    top=Side(border_style='thin'),
                                    bottom=Side(border_style='thin'))
                    cell.border = border
        # 调整每列的宽度
        for column in sheet.columns:
            max_length = 0
            column = list(column)  # 新建临时数组，用于计算最大宽度
            if column[0].row != 1:  # 排除第一行标题
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 1) * 1.2  # 加 2 个字符宽度，并乘以一个调整系数
                col_letter = get_column_letter(column[0].column)
                sheet.column_dimensions[col_letter].width = adjusted_width

        # 调整每行的高度
        for row in sheet.iter_rows():
            max_height = 0
            if row[0].row != 1:  # 排除第一行标题
                for cell in row:
                    try:
                        if len(str(cell.value)) > max_height:
                            max_height = len(cell.value)
                    except:
                        pass
                adjusted_height = (max_height + 1) * 0.4  # 加 2 个字符高度，并乘以一个调整系数
                row_num = row[0].row
                sheet.row_dimensions[row_num].height = adjusted_height
        writer.save()
        
        # 添加邮件附件
        mail.Attachments.Add(output_path)
        log(f"导出第{i+1}个Excel文件：{filename}，已完成")
        writer.close()
         
       # 发送邮件
        mail.Send()
        log(f"已发送第{i+1}封邮件至 {recipient}")
       
        # 如果不是最后一个收件人，则清空邮件收件人、主题，继续发送下一封邮件
        if i != len(recipient_list) - 1:
            mail = outlook.CreateItem(0)
            mail.To = ""           
            mail.Subject = ""


# In[ ]:




